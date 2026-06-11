"""Taskboard API — multi-user state, attachments, backups; SQLite + content-addressed file store."""
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import tempfile
import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path

from auth import (
    current_user_id,
    current_user_role,
    hash_password,
    login_user,
    logout_user,
    password_problems,
    require_admin,
    require_user,
    verify_password,
)
from flask import Flask, abort, jsonify, request, send_file
from migrate import migrate

DB_PATH = Path(os.environ.get("TASKBOARD_DB_PATH", "/var/lib/taskboard/data.db"))
FILES_DIR = Path(os.environ.get("TASKBOARD_FILES_DIR", "/var/lib/taskboard/files"))
MAX_UPLOAD_BYTES = int(os.environ.get("TASKBOARD_MAX_UPLOAD_BYTES", 25 * 1024 * 1024))
SECRET_KEY = os.environ.get("TASKBOARD_SECRET_KEY") or secrets.token_hex(32)
SESSION_DAYS = int(os.environ.get("TASKBOARD_SESSION_DAYS", "30"))
STATIC_DIR = os.environ.get("TASKBOARD_STATIC_DIR")  # if set, Flask also serves the frontend

APP = Flask(__name__)
APP.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES + 1024
APP.config["SECRET_KEY"] = SECRET_KEY
APP.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=SESSION_DAYS)
APP.config["SESSION_COOKIE_HTTPONLY"] = True
APP.config["SESSION_COOKIE_SAMESITE"] = "Lax"


def db():
    conn = sqlite3.connect(DB_PATH, isolation_level=None)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    (FILES_DIR / "tmp").mkdir(parents=True, exist_ok=True)
    with db() as conn:
        migrate(conn)


def now_iso() -> str:
    return datetime.now(UTC).isoformat()


def short_id() -> str:
    return secrets.token_urlsafe(8)


def file_path_for(sha: str) -> Path:
    return FILES_DIR / sha[:2] / sha


def user_dict(row):
    return {
        "id": row[0], "username": row[1], "role": row[2],
        "is_active": bool(row[3]), "created_at": row[4], "last_login_at": row[5],
    }


# ---------- session / auth ----------

@APP.get("/api/health")
def health():
    return jsonify(status="ok", ts=now_iso())


@APP.post("/api/session")
def login():
    payload = request.get_json(silent=True) or {}
    username = (payload.get("username") or "").strip().lower()
    password = payload.get("password") or ""
    if not username or not password:
        return jsonify(error="username and password required"), 400
    with db() as conn:
        row = conn.execute(
            "SELECT id, password_hash, role, is_active, totp_enabled FROM users WHERE username=?",
            (username,),
        ).fetchone()
    if not row or not row[3] or not verify_password(password, row[1]):
        return jsonify(error="invalid credentials"), 401
    uid, _, role, _, totp_enabled = row
    if totp_enabled:
        # Stash a pending login. Cleared by TOTP step or by next /api/session POST.
        from flask import session
        session.clear()
        session["pending_uid"] = uid
        session.permanent = False
        return jsonify(requires_totp=True, username=username), 200
    login_user(uid, role)
    with db() as conn:
        conn.execute("UPDATE users SET last_login_at=? WHERE id=?", (now_iso(), uid))
    return jsonify(id=uid, username=username, role=role)


@APP.post("/api/session/totp")
def login_totp():
    """Second step of login when 2FA is enabled. Consumes session['pending_uid']."""
    import pyotp
    from flask import session
    pending = session.get("pending_uid")
    if not pending:
        return jsonify(error="no pending login"), 400
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip()
    if not code or not code.isdigit() or len(code) not in (6, 8):
        return jsonify(error="invalid code format"), 400
    with db() as conn:
        row = conn.execute(
            "SELECT username, role, is_active, totp_secret, totp_enabled FROM users WHERE id=?",
            (pending,),
        ).fetchone()
    if not row or not row[2] or not row[4] or not row[3]:
        session.clear()
        return jsonify(error="invalid"), 401
    username, role, _, secret, _ = row
    if not pyotp.TOTP(secret).verify(code, valid_window=1):
        return jsonify(error="invalid code"), 401
    login_user(pending, role)
    with db() as conn:
        conn.execute("UPDATE users SET last_login_at=? WHERE id=?", (now_iso(), pending))
    return jsonify(id=pending, username=username, role=role)


@APP.delete("/api/session")
def logout():
    logout_user()
    return jsonify(ok=True)


@APP.get("/api/session")
def whoami():
    uid = current_user_id()
    if uid is None:
        return jsonify(authenticated=False), 200
    with db() as conn:
        row = conn.execute(
            "SELECT id, username, role, is_active FROM users WHERE id=?", (uid,)
        ).fetchone()
    if not row or not row[3]:
        logout_user()
        return jsonify(authenticated=False), 200
    return jsonify(authenticated=True, id=row[0], username=row[1], role=row[2])


# ---------- state ----------

@APP.get("/api/data")
@require_user
def get_data():
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT data, updated_at FROM state WHERE user_id=?", (uid,)
        ).fetchone()
    if not row:
        return jsonify(data=None, updated_at=None)
    return jsonify(data=json.loads(row[0]), updated_at=row[1])


@APP.post("/api/data")
@require_user
def put_data():
    payload = request.get_json(silent=True)
    if payload is None or "data" not in payload:
        return jsonify(error="missing 'data'"), 400
    body = json.dumps(payload["data"], separators=(",", ":"))
    if len(body) > 5_000_000:
        return jsonify(error="state too large"), 413
    uid = current_user_id()
    ts = now_iso()
    with db() as conn:
        conn.execute(
            "INSERT INTO state(user_id,data,updated_at) VALUES(?,?,?) "
            "ON CONFLICT(user_id) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at",
            (uid, body, ts),
        )
    return jsonify(ok=True, updated_at=ts)


# ---------- attachments ----------

@APP.get("/api/attachments")
@require_user
def list_attachments():
    task_id = request.args.get("task_id")
    if not task_id:
        return jsonify(error="missing task_id"), 400
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT id,task_id,filename,mime,size,sha256,uploaded_at "
            "FROM attachments WHERE task_id=? AND user_id=? ORDER BY uploaded_at DESC",
            (task_id, uid),
        ).fetchall()
    return jsonify([
        dict(zip(["id", "task_id", "filename", "mime", "size", "sha256", "uploaded_at"], r, strict=True))
        for r in rows
    ])


@APP.post("/api/attachments")
@require_user
def upload_attachment():
    task_id = request.args.get("task_id") or request.form.get("task_id")
    if not task_id:
        return jsonify(error="missing task_id"), 400
    f = request.files.get("file")
    if f is None or not f.filename:
        return jsonify(error="missing file"), 400

    h = hashlib.sha256()
    size = 0
    tmp_fd, tmp_path = tempfile.mkstemp(dir=FILES_DIR / "tmp")
    try:
        with os.fdopen(tmp_fd, "wb") as out:
            while True:
                chunk = f.stream.read(65536)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_UPLOAD_BYTES:
                    return jsonify(error="file exceeds limit"), 413
                h.update(chunk)
                out.write(chunk)
        if size == 0:
            return jsonify(error="empty file"), 400
        sha = h.hexdigest()
        dest = file_path_for(sha)
        dest.parent.mkdir(parents=True, exist_ok=True)
        if not dest.exists():
            os.replace(tmp_path, dest)
            tmp_path = None
        att_id = short_id()
        mime = f.mimetype or "application/octet-stream"
        ts = now_iso()
        uid = current_user_id()
        with db() as conn:
            conn.execute(
                "INSERT INTO attachments(id,user_id,task_id,filename,mime,size,sha256,uploaded_at) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (att_id, uid, task_id, f.filename, mime, size, sha, ts),
            )
        return jsonify(id=att_id, task_id=task_id, filename=f.filename,
                       mime=mime, size=size, sha256=sha, uploaded_at=ts), 201
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@APP.get("/api/attachments/<att_id>")
@require_user
def download_attachment(att_id):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT filename,mime,sha256,user_id FROM attachments WHERE id=?", (att_id,)
        ).fetchone()
    if not row:
        abort(404)
    filename, mime, sha, owner = row
    if owner != uid and current_user_role() != "admin":
        abort(403)
    path = file_path_for(sha)
    if not path.exists():
        abort(410)
    return send_file(path, mimetype=mime, as_attachment=False, download_name=filename)


@APP.delete("/api/attachments/<att_id>")
@require_user
def delete_attachment(att_id):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute("SELECT sha256, user_id FROM attachments WHERE id=?", (att_id,)).fetchone()
        if not row:
            return jsonify(error="not found"), 404
        sha, owner = row
        if owner != uid and current_user_role() != "admin":
            return jsonify(error="forbidden"), 403
        conn.execute("DELETE FROM attachments WHERE id=?", (att_id,))
        still = conn.execute("SELECT 1 FROM attachments WHERE sha256=? LIMIT 1", (sha,)).fetchone()
    if not still:
        path = file_path_for(sha)
        if path.exists():
            path.unlink()
    return jsonify(ok=True)


# ---------- backups (per-user) ----------

@APP.get("/api/backups")
@require_user
def list_backups():
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT id,size,source,note,created_at FROM backups "
            "WHERE user_id=? ORDER BY id DESC LIMIT 200", (uid,)
        ).fetchall()
    return jsonify([
        dict(zip(["id", "size", "source", "note", "created_at"], r, strict=True)) for r in rows
    ])


@APP.post("/api/backups")
@require_user
def create_backup():
    payload = request.get_json(silent=True) or {}
    source = (payload.get("source") or "manual")[:32]
    note = payload.get("note") or None
    data = payload.get("data")
    uid = current_user_id()
    if data is None:
        with db() as conn:
            row = conn.execute("SELECT data FROM state WHERE user_id=?", (uid,)).fetchone()
        if not row:
            return jsonify(error="no current state to snapshot"), 400
        body = row[0]
    else:
        body = json.dumps(data, separators=(",", ":"))
    ts = now_iso()
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO backups(user_id,data,size,source,note,created_at) VALUES(?,?,?,?,?,?)",
            (uid, body, len(body), source, note, ts),
        )
        bid = cur.lastrowid
    return jsonify(id=bid, size=len(body), source=source, note=note, created_at=ts), 201


@APP.get("/api/backups/<int:bid>")
@require_user
def get_backup(bid):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT id,data,size,source,note,created_at,user_id FROM backups WHERE id=?", (bid,)
        ).fetchone()
    if not row:
        abort(404)
    if row[6] != uid and current_user_role() != "admin":
        abort(403)
    return jsonify(
        id=row[0], data=json.loads(row[1]), size=row[2],
        source=row[3], note=row[4], created_at=row[5],
    )


@APP.delete("/api/backups/<int:bid>")
@require_user
def delete_backup(bid):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute("SELECT user_id FROM backups WHERE id=?", (bid,)).fetchone()
        if not row:
            return jsonify(error="not found"), 404
        if row[0] != uid and current_user_role() != "admin":
            return jsonify(error="forbidden"), 403
        conn.execute("DELETE FROM backups WHERE id=?", (bid,))
    return jsonify(ok=True)


# ---------- notes (per-user daily capture pad + freeform notebooks) ----------

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ID_RE = re.compile(r"^[A-Za-z0-9_-]{6,32}$")  # client-supplied note ids (offline create)
_TAG_RE = re.compile(r"(?<![\w/])#([A-Za-z0-9][\w-]{0,63})")
_WIKILINK_RE = re.compile(r"\[\[([^\[\]\n]{1,200})\]\]")


def _today_iso() -> str:
    return datetime.now(UTC).date().isoformat()


def _note_dict(row):
    # row order: id,date,title,body,created_at,updated_at,pinned,sort_order,notebook_id,project_id
    return {
        "id": row[0], "date": row[1], "title": row[2], "body": row[3],
        "created_at": row[4], "updated_at": row[5],
        "pinned": bool(row[6]) if len(row) > 6 else False,
        "sort_order": row[7] if len(row) > 7 else 0,
        "notebook_id": row[8] if len(row) > 8 else None,
        "project_id": row[9] if len(row) > 9 else None,
        "task_id": row[10] if len(row) > 10 else None,
    }


_NOTE_COLS = "id,date,title,body,created_at,updated_at,pinned,sort_order,notebook_id,project_id,task_id"


def _extract_tags(body: str) -> list[str]:
    if not body:
        return []
    seen, out = set(), []
    for m in _TAG_RE.finditer(body):
        t = m.group(1).lower()
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


@APP.get("/api/notes")
@require_user
def list_notes():
    """Three modes:
       - default: notes for ?date=YYYY-MM-DD (today if omitted) where notebook_id IS NULL
       - ?notebook_id=X: all notes in that notebook
       - ?all=1: all notes, newest-updated first (paginated by ?limit=&offset=)
       - ?pinned=1: only pinned notes
       - ?tag=foo: notes whose body contains #foo (case-insensitive)
       - ?q=...: brute-force title/body match (lightweight)
    """
    uid = current_user_id()
    notebook_id = request.args.get("notebook_id")
    project_id = request.args.get("project_id")
    task_id = request.args.get("task_id")
    all_flag = request.args.get("all") == "1"
    pinned_flag = request.args.get("pinned") == "1"
    tag = request.args.get("tag")
    q = request.args.get("q")
    limit = max(1, min(int(request.args.get("limit") or 500), 500))
    offset = max(0, int(request.args.get("offset") or 0))

    where = ["user_id=?"]
    params: list = [uid]
    order = "ORDER BY pinned DESC, sort_order, created_at"

    if all_flag:
        order = "ORDER BY pinned DESC, updated_at DESC"
    elif pinned_flag:
        where.append("pinned=1")
        order = "ORDER BY updated_at DESC"
    elif notebook_id:
        where.append("notebook_id=?")
        params.append(notebook_id)
    elif task_id:
        where.append("task_id=?")
        params.append(task_id)
        order = "ORDER BY pinned DESC, updated_at DESC"
    elif project_id:
        where.append("project_id=?")
        params.append(project_id)
        order = "ORDER BY pinned DESC, updated_at DESC"
    elif tag:
        # We can't index this — full-scan + Python filter (capped by date filter if any)
        where.append("notebook_id IS NULL OR notebook_id IS NOT NULL")  # noop, tag filter applied below
    else:
        date = request.args.get("date") or _today_iso()
        if not _DATE_RE.match(date):
            return jsonify(error="date must be YYYY-MM-DD"), 400
        where.append("date=?")
        where.append("notebook_id IS NULL")
        params.append(date)

    if q:
        ql = f"%{q.lower()}%"
        where.append("(LOWER(title) LIKE ? OR LOWER(body) LIKE ?)")
        params += [ql, ql]

    sql = f"SELECT {_NOTE_COLS} FROM notes WHERE " + " AND ".join(where) + f" {order} LIMIT ? OFFSET ?"
    params += [limit, offset]
    with db() as conn:
        rows = conn.execute(sql, params).fetchall()

    notes = [_note_dict(r) for r in rows]
    if tag:
        tlow = tag.lower()
        notes = [n for n in notes if tlow in {t for t in _extract_tags(n["body"])}]
    return jsonify(notes)


@APP.get("/api/notes/dates")
@require_user
def list_note_dates():
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT date, COUNT(*) FROM notes WHERE user_id=? AND notebook_id IS NULL "
            "GROUP BY date ORDER BY date DESC LIMIT 365",
            (uid,),
        ).fetchall()
    return jsonify([{"date": r[0], "count": r[1]} for r in rows])


@APP.get("/api/notes/tags")
@require_user
def list_tags():
    """Distinct tags across this user's notes, sorted by count desc."""
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT body FROM notes WHERE user_id=?", (uid,)
        ).fetchall()
    counts: dict[str, int] = {}
    for (body,) in rows:
        for t in _extract_tags(body or ""):
            counts[t] = counts.get(t, 0) + 1
    return jsonify(sorted(
        [{"tag": k, "count": v} for k, v in counts.items()],
        key=lambda r: (-r["count"], r["tag"]),
    ))


@APP.get("/api/notes/<note_id>")
@require_user
def get_note(note_id):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE id=? AND user_id=?",
            (note_id, uid),
        ).fetchone()
    if not row:
        return jsonify(error="not found"), 404
    return jsonify(_note_dict(row))


@APP.get("/api/notes/<note_id>/backlinks")
@require_user
def note_backlinks(note_id):
    """Notes whose body contains [[<this note's title>]]."""
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT title FROM notes WHERE id=? AND user_id=?", (note_id, uid)
        ).fetchone()
        if not row:
            return jsonify(error="not found"), 404
        title = (row[0] or "").strip()
        if not title:
            return jsonify([])
        # Case-insensitive substring match for [[title]] — Python filters precise hits.
        candidate = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE user_id=? AND id<>? AND LOWER(body) LIKE ?",
            (uid, note_id, f"%[[{title.lower()}%"),
        ).fetchall()
    out = []
    tlow = title.lower()
    for r in candidate:
        body = r[3] or ""
        for m in _WIKILINK_RE.finditer(body):
            if m.group(1).strip().lower() == tlow:
                out.append(_note_dict(r))
                break
    return jsonify(out)


@APP.get("/api/notes/by-title")
@require_user
def find_note_by_title():
    """Resolve a wiki-link target. Case-insensitive exact title match; first hit wins."""
    title = (request.args.get("title") or "").strip()
    if not title:
        return jsonify(error="missing title"), 400
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE user_id=? AND LOWER(title)=? "
            "ORDER BY updated_at DESC LIMIT 1",
            (uid, title.lower()),
        ).fetchone()
    if not row:
        return jsonify(error="not found"), 404
    return jsonify(_note_dict(row))


@APP.get("/api/notes/titles")
@require_user
def list_note_titles():
    """For wiki-link autocomplete: lightweight list of titles."""
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT id,title,date FROM notes WHERE user_id=? AND title<>'' "
            "ORDER BY updated_at DESC LIMIT 500",
            (uid,),
        ).fetchall()
    return jsonify([{"id": r[0], "title": r[1], "date": r[2]} for r in rows])


@APP.post("/api/notes")
@require_user
def create_note():
    payload = request.get_json(silent=True) or {}
    notebook_id = payload.get("notebook_id")
    if notebook_id == "":
        notebook_id = None
    uid = current_user_id()

    if notebook_id:
        # Validate notebook ownership
        with db() as conn:
            row = conn.execute(
                "SELECT 1 FROM notebooks WHERE id=? AND user_id=?", (notebook_id, uid)
            ).fetchone()
        if not row:
            return jsonify(error="notebook not found"), 404
        date = payload.get("date") or _today_iso()
    else:
        date = payload.get("date") or _today_iso()
    if not _DATE_RE.match(date):
        return jsonify(error="date must be YYYY-MM-DD"), 400
    title = (payload.get("title") or "")[:300]
    body = (payload.get("body") or "")[:50_000]
    pinned = 1 if payload.get("pinned") else 0
    # Optional client-supplied id enables idempotent offline create: a note
    # drafted while offline gets a stable id on the device, so replaying the
    # queued create (or a double-submit) returns the existing note instead of
    # duplicating it.
    cid = payload.get("id")
    if cid is not None:
        if not _ID_RE.match(str(cid)):
            return jsonify(error="invalid id"), 400
        cid = str(cid)
        with db() as conn:
            existing = conn.execute(
                f"SELECT {_NOTE_COLS} FROM notes WHERE id=? AND user_id=?", (cid, uid)
            ).fetchone()
        if existing:
            return jsonify(_note_dict(existing)), 200
        nid = cid
    else:
        nid = short_id()
    ts = now_iso()
    with db() as conn:
        # Place at end of current bucket (highest sort_order + 1)
        if notebook_id:
            row = conn.execute(
                "SELECT COALESCE(MAX(sort_order),0)+1 FROM notes WHERE user_id=? AND notebook_id=?",
                (uid, notebook_id),
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT COALESCE(MAX(sort_order),0)+1 FROM notes WHERE user_id=? AND date=? AND notebook_id IS NULL",
                (uid, date),
            ).fetchone()
        sort_order = row[0] if row else 0
        conn.execute(
            "INSERT INTO notes(id,user_id,date,title,body,created_at,updated_at,pinned,sort_order,notebook_id,project_id,task_id) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (nid, uid, date, title, body, ts, ts, pinned, sort_order, notebook_id,
             payload.get("project_id") or None, payload.get("task_id") or None),
        )
        row = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE id=?", (nid,)
        ).fetchone()
    return jsonify(_note_dict(row)), 201


@APP.patch("/api/notes/<note_id>")
@require_user
def update_note(note_id):
    payload = request.get_json(silent=True) or {}
    sets, params = [], []
    if "title" in payload:
        sets.append("title=?")
        params.append(str(payload["title"])[:300])
    if "body" in payload:
        sets.append("body=?")
        params.append(str(payload["body"])[:50_000])
    if "date" in payload:
        if not _DATE_RE.match(payload["date"]):
            return jsonify(error="date must be YYYY-MM-DD"), 400
        sets.append("date=?")
        params.append(payload["date"])
    if "pinned" in payload:
        sets.append("pinned=?")
        params.append(1 if payload["pinned"] else 0)
    if "notebook_id" in payload:
        nb = payload["notebook_id"]
        if nb in ("", None):
            sets.append("notebook_id=NULL")
        else:
            uid = current_user_id()
            with db() as conn:
                ok = conn.execute(
                    "SELECT 1 FROM notebooks WHERE id=? AND user_id=?", (nb, uid)
                ).fetchone()
            if not ok:
                return jsonify(error="notebook not found"), 404
            sets.append("notebook_id=?")
            params.append(nb)
    if "project_id" in payload:
        pid = payload["project_id"]
        if pid in ("", None):
            sets.append("project_id=NULL")
        else:
            sets.append("project_id=?")
            params.append(str(pid)[:64])
    if "task_id" in payload:
        tid = payload["task_id"]
        if tid in ("", None):
            sets.append("task_id=NULL")
        else:
            sets.append("task_id=?")
            params.append(str(tid)[:64])
    if not sets:
        return jsonify(error="no fields to update"), 400
    sets.append("updated_at=?")
    params.append(now_iso())
    uid = current_user_id()
    params.extend([note_id, uid])
    with db() as conn:
        cur = conn.execute(
            f"UPDATE notes SET {', '.join(sets)} WHERE id=? AND user_id=?", params
        )
        if cur.rowcount == 0:
            return jsonify(error="not found"), 404
        row = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE id=?", (note_id,)
        ).fetchone()
    return jsonify(_note_dict(row))


@APP.delete("/api/notes/<note_id>")
@require_user
def delete_note(note_id):
    uid = current_user_id()
    with db() as conn:
        cur = conn.execute("DELETE FROM notes WHERE id=? AND user_id=?", (note_id, uid))
        if cur.rowcount == 0:
            return jsonify(error="not found"), 404
    return jsonify(ok=True)


@APP.post("/api/notes/reorder")
@require_user
def reorder_notes():
    """Body: {ids: ["a","b","c"]} — assigns sort_order = index in that array.
    All notes must belong to the caller; otherwise 404."""
    payload = request.get_json(silent=True) or {}
    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not all(isinstance(x, str) for x in ids):
        return jsonify(error="ids must be a list of strings"), 400
    uid = current_user_id()
    ts = now_iso()
    with db() as conn:
        for i, nid in enumerate(ids):
            cur = conn.execute(
                "UPDATE notes SET sort_order=?, updated_at=? WHERE id=? AND user_id=?",
                (i, ts, nid, uid),
            )
            if cur.rowcount == 0:
                return jsonify(error=f"note {nid} not found"), 404
    return jsonify(ok=True, count=len(ids))


@APP.post("/api/notes/<note_id>/duplicate")
@require_user
def duplicate_note(note_id):
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            f"SELECT {_NOTE_COLS} FROM notes WHERE id=? AND user_id=?",
            (note_id, uid),
        ).fetchone()
        if not row:
            return jsonify(error="not found"), 404
        nid = short_id()
        ts = now_iso()
        new_title = (row[2] or "Untitled") + " (copy)"
        conn.execute(
            "INSERT INTO notes(id,user_id,date,title,body,created_at,updated_at,pinned,sort_order,notebook_id,project_id,task_id) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (nid, uid, row[1], new_title, row[3], ts, ts, 0, (row[7] or 0) + 1, row[8],
             row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None),
        )
        new_row = conn.execute(f"SELECT {_NOTE_COLS} FROM notes WHERE id=?", (nid,)).fetchone()
    return jsonify(_note_dict(new_row)), 201


# ---------- notebooks ----------

def _notebook_dict(row, count=None):
    out = {
        "id": row[0], "name": row[1], "color": row[2], "icon": row[3],
        "sort_order": row[4], "created_at": row[5],
    }
    if count is not None:
        out["count"] = count
    return out


@APP.get("/api/notebooks")
@require_user
def list_notebooks():
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT n.id,n.name,n.color,n.icon,n.sort_order,n.created_at,"
            "(SELECT COUNT(*) FROM notes WHERE notebook_id=n.id) "
            "FROM notebooks n WHERE n.user_id=? ORDER BY n.sort_order,n.created_at",
            (uid,),
        ).fetchall()
    return jsonify([_notebook_dict(r[:6], r[6]) for r in rows])


@APP.post("/api/notebooks")
@require_user
def create_notebook():
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()[:120]
    if not name:
        return jsonify(error="name required"), 400
    color = (payload.get("color") or "#c1542a")[:32]
    icon = (payload.get("icon") or "book")[:32]
    nid = short_id()
    ts = now_iso()
    uid = current_user_id()
    with db() as conn:
        sort_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order),0)+1 FROM notebooks WHERE user_id=?", (uid,)
        ).fetchone()[0]
        conn.execute(
            "INSERT INTO notebooks(id,user_id,name,color,icon,sort_order,created_at) "
            "VALUES(?,?,?,?,?,?,?)",
            (nid, uid, name, color, icon, sort_order, ts),
        )
    return jsonify(_notebook_dict((nid, name, color, icon, sort_order, ts), 0)), 201


@APP.patch("/api/notebooks/<nb_id>")
@require_user
def update_notebook(nb_id):
    payload = request.get_json(silent=True) or {}
    sets, params = [], []
    if "name" in payload:
        n = (payload["name"] or "").strip()[:120]
        if not n:
            return jsonify(error="name required"), 400
        sets.append("name=?")
        params.append(n)
    if "color" in payload:
        sets.append("color=?")
        params.append(str(payload["color"])[:32])
    if "icon" in payload:
        sets.append("icon=?")
        params.append(str(payload["icon"])[:32])
    if "sort_order" in payload:
        sets.append("sort_order=?")
        params.append(int(payload["sort_order"]))
    if not sets:
        return jsonify(error="no fields to update"), 400
    uid = current_user_id()
    params.extend([nb_id, uid])
    with db() as conn:
        cur = conn.execute(
            f"UPDATE notebooks SET {', '.join(sets)} WHERE id=? AND user_id=?", params
        )
        if cur.rowcount == 0:
            return jsonify(error="not found"), 404
        row = conn.execute(
            "SELECT id,name,color,icon,sort_order,created_at FROM notebooks WHERE id=?", (nb_id,)
        ).fetchone()
    return jsonify(_notebook_dict(row))


@APP.delete("/api/notebooks/<nb_id>")
@require_user
def delete_notebook(nb_id):
    """Deleting a notebook un-files its notes back to NULL — does NOT delete them."""
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT 1 FROM notebooks WHERE id=? AND user_id=?", (nb_id, uid)
        ).fetchone()
        if not row:
            return jsonify(error="not found"), 404
        conn.execute(
            "UPDATE notes SET notebook_id=NULL WHERE notebook_id=? AND user_id=?",
            (nb_id, uid),
        )
        conn.execute("DELETE FROM notebooks WHERE id=? AND user_id=?", (nb_id, uid))
    return jsonify(ok=True)


# ---------- note image upload (reuses the content-addressed blob store) ----------

@APP.post("/api/notes/<note_id>/images")
@require_user
def upload_note_image(note_id):
    """Multipart upload an image to be embedded in a note. Returns {url, id}."""
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT 1 FROM notes WHERE id=? AND user_id=?", (note_id, uid)
        ).fetchone()
    if not row:
        return jsonify(error="not found"), 404
    f = request.files.get("file")
    if f is None or not f.filename:
        return jsonify(error="missing file"), 400
    mime = f.mimetype or "application/octet-stream"
    if not mime.startswith("image/"):
        return jsonify(error="must be image/*"), 400

    h = hashlib.sha256()
    size = 0
    tmp_fd, tmp_path = tempfile.mkstemp(dir=FILES_DIR / "tmp")
    try:
        with os.fdopen(tmp_fd, "wb") as out:
            while True:
                chunk = f.stream.read(65536)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_UPLOAD_BYTES:
                    return jsonify(error="file exceeds limit"), 413
                h.update(chunk)
                out.write(chunk)
        if size == 0:
            return jsonify(error="empty file"), 400
        sha = h.hexdigest()
        dest = file_path_for(sha)
        dest.parent.mkdir(parents=True, exist_ok=True)
        if not dest.exists():
            os.replace(tmp_path, dest)
            tmp_path = None
        # Optional client-supplied attachment id: a sketch/photo drawn while
        # offline gets a stable id and markdown URL on the device, so replaying
        # the queued upload is idempotent and the embed resolves once synced.
        cid = request.form.get("id")
        if cid:
            if not _ID_RE.match(cid):
                return jsonify(error="invalid id"), 400
            with db() as conn:
                existing = conn.execute(
                    "SELECT filename,mime,size FROM attachments WHERE id=? AND user_id=?",
                    (cid, uid),
                ).fetchone()
            if existing:
                return jsonify(id=cid, url=f"/api/attachments/{cid}",
                               filename=existing[0], mime=existing[1], size=existing[2]), 200
            att_id = cid
        else:
            att_id = short_id()
        ts = now_iso()
        with db() as conn:
            conn.execute(
                "INSERT INTO attachments(id,user_id,task_id,filename,mime,size,sha256,uploaded_at) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (att_id, uid, f"note:{note_id}", f.filename, mime, size, sha, ts),
            )
        return jsonify(id=att_id, url=f"/api/attachments/{att_id}",
                       filename=f.filename, mime=mime, size=size), 201
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _markdown_for_day(date: str, rows) -> str:
    """Render a list of (id,title,body,created_at,updated_at) note rows for one day."""
    out = [f"# Notes — {date}", ""]
    if not rows:
        out.append("_(no notes)_")
        return "\n".join(out) + "\n"
    for i, (_nid, title, body, created_at, _updated_at) in enumerate(rows):
        if i > 0:
            out.append("---")
            out.append("")
        out.append(f"## {title or 'Untitled'}")
        out.append(f"*created {created_at}*")
        out.append("")
        if body:
            out.append(body.rstrip())
        out.append("")
    return "\n".join(out) + "\n"


@APP.get("/api/notes/export")
@require_user
def export_notes_day():
    date = request.args.get("date") or _today_iso()
    if not _DATE_RE.match(date):
        return jsonify(error="date must be YYYY-MM-DD"), 400
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT id,title,body,created_at,updated_at FROM notes "
            "WHERE user_id=? AND date=? ORDER BY created_at",
            (uid, date),
        ).fetchall()
    body = _markdown_for_day(date, rows)
    return body, 200, {
        "Content-Type": "text/markdown; charset=utf-8",
        "Content-Disposition": f'attachment; filename="notes-{date}.md"',
    }


@APP.get("/api/notes/export.zip")
@require_user
def export_notes_zip():
    uid = current_user_id()
    with db() as conn:
        rows = conn.execute(
            "SELECT date, id, title, body, created_at, updated_at FROM notes "
            "WHERE user_id=? ORDER BY date, created_at",
            (uid,),
        ).fetchall()
    by_date: dict[str, list] = {}
    for date, nid, title, body, created_at, updated_at in rows:
        by_date.setdefault(date, []).append((nid, title, body, created_at, updated_at))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Always include a README so an empty zip isn't broken
        zf.writestr(
            "README.md",
            f"# Taskboard notes export\n\n"
            f"Exported {now_iso()}.\n"
            f"Files: one per date, format YYYY-MM-DD.md.\n"
            f"Total dates: {len(by_date)}.\n",
        )
        for date, day_rows in by_date.items():
            zf.writestr(f"notes/{date}.md", _markdown_for_day(date, day_rows))
    buf.seek(0)
    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    return buf.getvalue(), 200, {
        "Content-Type": "application/zip",
        "Content-Disposition": f'attachment; filename="taskboard-notes-{stamp}.zip"',
    }


@APP.post("/api/export/xlsx")
@require_user
def export_xlsx():
    """Generic spreadsheet export. Body: {filename, sheets:[{name, headers:[...],
    rows:[[...]]}]}. Kept schema-agnostic so the client (which owns the board
    blob) decides what to put in each sheet — projects, tasks, port maps, etc."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return jsonify(error="xlsx export not available on server"), 501

    payload = request.get_json(silent=True) or {}
    sheets = payload.get("sheets") or [{"name": "Sheet1", "headers": [], "rows": []}]
    fn = re.sub(r"[^A-Za-z0-9_.-]", "_", str(payload.get("filename") or "export"))[:80] or "export"
    if not fn.endswith(".xlsx"):
        fn += ".xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for i, sh in enumerate(sheets[:25]):
        raw = re.sub(r"[\[\]:*?/\\]", " ", str(sh.get("name") or f"Sheet{i + 1}"))[:31].strip() or f"Sheet{i + 1}"
        name, n = raw, 2
        while name.lower() in used:
            suffix = f" ({n})"; name = raw[:31 - len(suffix)] + suffix; n += 1
        used.add(name.lower())
        ws = wb.create_sheet(title=name)
        headers = sh.get("headers") or []
        if headers:
            ws.append([str(h) for h in headers][:50])
            for c in ws[1]:
                c.font = Font(bold=True)
        for r in (sh.get("rows") or [])[:10000]:
            ws.append([("" if v is None else v if isinstance(v, (int, float)) else str(v)) for v in r][:50])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fn,
    )


# ---------- 2FA (TOTP) setup / enable / disable ----------

@APP.post("/api/2fa/setup")
@require_user
def totp_setup():
    """Generate a candidate secret. Stores it but DOES NOT enable until /enable."""
    import pyotp
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT username, totp_enabled FROM users WHERE id=?", (uid,)
        ).fetchone()
    if not row:
        abort(404)
    if row[1]:
        return jsonify(error="2FA already enabled — disable first to re-enroll"), 400
    secret = pyotp.random_base32()
    with db() as conn:
        conn.execute(
            "UPDATE users SET totp_secret=?, totp_enabled=0 WHERE id=?",
            (secret, uid),
        )
    uri = pyotp.TOTP(secret).provisioning_uri(name=row[0], issuer_name="Taskboard")
    return jsonify(secret=secret, uri=uri)


@APP.post("/api/2fa/enable")
@require_user
def totp_enable():
    """Verify a code against the candidate secret, then flip enabled=1."""
    import pyotp
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip()
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT totp_secret, totp_enabled FROM users WHERE id=?", (uid,)
        ).fetchone()
    if not row or not row[0]:
        return jsonify(error="run setup first"), 400
    if row[1]:
        return jsonify(error="already enabled"), 400
    if not pyotp.TOTP(row[0]).verify(code, valid_window=1):
        return jsonify(error="invalid code"), 401
    with db() as conn:
        conn.execute("UPDATE users SET totp_enabled=1 WHERE id=?", (uid,))
    return jsonify(ok=True, enabled=True)


@APP.post("/api/2fa/disable")
@require_user
def totp_disable():
    """Disable 2FA. Requires a current TOTP code as proof of possession."""
    import pyotp
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip()
    uid = current_user_id()
    with db() as conn:
        row = conn.execute(
            "SELECT totp_secret, totp_enabled FROM users WHERE id=?", (uid,)
        ).fetchone()
    if not row or not row[1]:
        return jsonify(error="2FA is not enabled"), 400
    if not pyotp.TOTP(row[0]).verify(code, valid_window=1):
        return jsonify(error="invalid code"), 401
    with db() as conn:
        conn.execute("UPDATE users SET totp_secret=NULL, totp_enabled=0 WHERE id=?", (uid,))
    return jsonify(ok=True, enabled=False)


@APP.get("/api/2fa/status")
@require_user
def totp_status():
    uid = current_user_id()
    with db() as conn:
        row = conn.execute("SELECT totp_enabled FROM users WHERE id=?", (uid,)).fetchone()
    return jsonify(enabled=bool(row and row[0]))


# ---------- iCal feed (no-auth subscription URL with per-user token) ----------

def _ical_escape(s: str) -> str:
    if not s:
        return ""
    return (s.replace("\\", "\\\\").replace(",", "\\,")
             .replace(";", "\\;").replace("\n", "\\n"))


def _ical_for_state(state: dict, username: str) -> str:
    """Render a user's open, due-dated tasks as a VCALENDAR."""
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        f"PRODID:-//taskboard//{username}//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:Taskboard — {_ical_escape(username)}",
    ]
    now = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    for p in state.get("projects", []):
        for t in p.get("tasks", []):
            if t.get("completed"):
                continue
            due = t.get("dueDate")
            if not due:
                continue
            try:
                d = datetime.strptime(due, "%Y-%m-%d").date()
            except ValueError:
                continue
            uid = f"{t.get('id','x')}@taskboard.local"
            summary = _ical_escape(t.get("title") or "Untitled")
            desc_parts = []
            if p.get("name"):
                desc_parts.append(f"Project: {p['name']}")
            if t.get("priority") and t["priority"] != "none":
                desc_parts.append(f"Priority: {t['priority']}")
            if t.get("description"):
                desc_parts.append(t["description"])
            description = _ical_escape("\n".join(desc_parts))
            lines += [
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{now}",
                f"DTSTART;VALUE=DATE:{d.strftime('%Y%m%d')}",
                f"DTEND;VALUE=DATE:{(d + timedelta(days=1)).strftime('%Y%m%d')}",
                f"SUMMARY:{summary}",
                f"DESCRIPTION:{description}",
                "TRANSP:TRANSPARENT",
                "END:VEVENT",
            ]
    lines.append("END:VCALENDAR")
    # iCal RFC 5545 wants CRLF line endings
    return "\r\n".join(lines) + "\r\n"


@APP.get("/api/feed/token")
@require_user
def get_feed_token():
    uid = current_user_id()
    with db() as conn:
        row = conn.execute("SELECT feed_token FROM users WHERE id=?", (uid,)).fetchone()
    token = row[0] if row else None
    if not token:
        token = secrets.token_urlsafe(24)
        with db() as conn:
            conn.execute("UPDATE users SET feed_token=? WHERE id=?", (token, uid))
    return jsonify(token=token, url=f"/api/calendar/{token}.ics")


@APP.post("/api/feed/token")
@require_user
def rotate_feed_token():
    uid = current_user_id()
    token = secrets.token_urlsafe(24)
    with db() as conn:
        conn.execute("UPDATE users SET feed_token=? WHERE id=?", (token, uid))
    return jsonify(token=token, url=f"/api/calendar/{token}.ics")


@APP.get("/api/calendar/<token>.ics")
def calendar_feed(token):
    if not token or len(token) < 16:
        abort(404)
    with db() as conn:
        row = conn.execute(
            "SELECT id, username FROM users WHERE feed_token=? AND is_active=1",
            (token,),
        ).fetchone()
        if not row:
            abort(404)
        uid, username = row
        state_row = conn.execute("SELECT data FROM state WHERE user_id=?", (uid,)).fetchone()
    state = json.loads(state_row[0]) if state_row else {}
    body = _ical_for_state(state, username)
    return body, 200, {
        "Content-Type": "text/calendar; charset=utf-8",
        "Cache-Control": "public, max-age=300",
    }


# ---------- search (across this user's projects/tasks/subtasks/notes) ----------

def _snippet(text: str, q: str, before: int = 40, after: int = 100) -> str:
    if not text:
        return ""
    low = text.lower()
    idx = low.find(q.lower())
    if idx < 0:
        return text[: before + after]
    start = max(0, idx - before)
    return ("…" if start > 0 else "") + text[start: idx + len(q) + after]


@APP.get("/api/search")
@require_user
def search():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify(results=[])
    qlow = q.lower()
    uid = current_user_id()
    results: list = []

    with db() as conn:
        state_row = conn.execute("SELECT data FROM state WHERE user_id=?", (uid,)).fetchone()
        note_rows = conn.execute(
            "SELECT id,date,title,body FROM notes WHERE user_id=? "
            "AND (LOWER(title) LIKE ? OR LOWER(body) LIKE ?) "
            "ORDER BY date DESC LIMIT 100",
            (uid, f"%{qlow}%", f"%{qlow}%"),
        ).fetchall()

    if state_row:
        try:
            state = json.loads(state_row[0])
        except json.JSONDecodeError:
            state = {}
        for p in state.get("projects", []):
            pname = p.get("name", "")
            if qlow in pname.lower():
                results.append({
                    "kind": "project", "id": p.get("id"),
                    "title": pname, "snippet": "",
                })
            for t in p.get("tasks", []):
                title = t.get("title", "") or ""
                desc = t.get("description", "") or ""
                if qlow in title.lower() or qlow in desc.lower():
                    results.append({
                        "kind": "task",
                        "id": t.get("id"),
                        "project_id": p.get("id"),
                        "project_name": pname,
                        "title": title or "Untitled",
                        "snippet": _snippet(desc, q),
                        "completed": bool(t.get("completed")),
                        "due_date": t.get("dueDate", ""),
                    })
                for st in t.get("subtasks", []):
                    sttitle = st.get("title", "") or ""
                    if qlow in sttitle.lower():
                        results.append({
                            "kind": "subtask",
                            "id": st.get("id"),
                            "task_id": t.get("id"),
                            "project_id": p.get("id"),
                            "project_name": pname,
                            "title": sttitle,
                            "parent_title": title or "Untitled",
                        })

    for r in note_rows:
        body = r[3] or ""
        title = r[2] or ""
        results.append({
            "kind": "note",
            "id": r[0],
            "date": r[1],
            "title": title or "Untitled note",
            "snippet": _snippet(body, q),
        })

    return jsonify(results=results[:200])


# ---------- users (admin) ----------

@APP.get("/api/users")
@require_admin
def list_users():
    with db() as conn:
        rows = conn.execute(
            "SELECT id, username, role, is_active, created_at, last_login_at FROM users ORDER BY id"
        ).fetchall()
    return jsonify([user_dict(r) for r in rows])


@APP.post("/api/users")
@require_admin
def create_user():
    payload = request.get_json(silent=True) or {}
    username = (payload.get("username") or "").strip().lower()
    password = payload.get("password") or ""
    role = payload.get("role") or "user"
    if not username or not username.replace("_", "").replace("-", "").isalnum() or len(username) > 32:
        return jsonify(error="invalid username (alphanumeric, underscore, hyphen; max 32)"), 400
    if role not in ("admin", "user"):
        return jsonify(error="role must be 'admin' or 'user'"), 400
    err = password_problems(password)
    if err:
        return jsonify(error=err), 400
    ph = hash_password(password)
    ts = now_iso()
    try:
        with db() as conn:
            cur = conn.execute(
                "INSERT INTO users(username,password_hash,role,is_active,created_at) VALUES(?,?,?,1,?)",
                (username, ph, role, ts),
            )
        uid = cur.lastrowid
    except sqlite3.IntegrityError:
        return jsonify(error="username already taken"), 409
    return jsonify(id=uid, username=username, role=role, is_active=True,
                   created_at=ts, last_login_at=None), 201


@APP.patch("/api/users/<int:uid>")
@require_admin
def update_user(uid):
    payload = request.get_json(silent=True) or {}
    sets = []
    params = []
    if "role" in payload:
        if payload["role"] not in ("admin", "user"):
            return jsonify(error="role must be 'admin' or 'user'"), 400
        sets.append("role=?")
        params.append(payload["role"])
    if "is_active" in payload:
        sets.append("is_active=?")
        params.append(1 if payload["is_active"] else 0)
    if "password" in payload:
        err = password_problems(payload["password"])
        if err:
            return jsonify(error=err), 400
        sets.append("password_hash=?")
        params.append(hash_password(payload["password"]))
    if not sets:
        return jsonify(error="no fields to update"), 400
    params.append(uid)
    with db() as conn:
        cur = conn.execute(f"UPDATE users SET {', '.join(sets)} WHERE id=?", params)
        if cur.rowcount == 0:
            return jsonify(error="not found"), 404
        row = conn.execute(
            "SELECT id, username, role, is_active, created_at, last_login_at FROM users WHERE id=?", (uid,)
        ).fetchone()
    return jsonify(user_dict(row))


@APP.delete("/api/users/<int:uid>")
@require_admin
def delete_user(uid):
    if uid == current_user_id():
        return jsonify(error="cannot delete yourself"), 400
    with db() as conn:
        # Check admin count to prevent removing last admin
        is_target_admin = conn.execute(
            "SELECT 1 FROM users WHERE id=? AND role='admin' AND is_active=1", (uid,)
        ).fetchone()
        if is_target_admin:
            n = conn.execute(
                "SELECT COUNT(*) FROM users WHERE role='admin' AND is_active=1"
            ).fetchone()[0]
            if n <= 1:
                return jsonify(error="cannot remove the last active admin"), 400
        # Cascade: delete the user's data
        conn.execute("DELETE FROM state WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM attachments WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM backups WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM notes WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM notebooks WHERE user_id=?", (uid,))
        cur = conn.execute("DELETE FROM users WHERE id=?", (uid,))
        if cur.rowcount == 0:
            return jsonify(error="not found"), 404
    return jsonify(ok=True)


@APP.get("/api/admin/stats")
@require_admin
def admin_stats():
    with db() as conn:
        users_total = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        users_active = conn.execute("SELECT COUNT(*) FROM users WHERE is_active=1").fetchone()[0]
        states = conn.execute("SELECT COUNT(*) FROM state").fetchone()[0]
        atts = conn.execute("SELECT COUNT(*), COALESCE(SUM(size),0) FROM attachments").fetchone()
        bks = conn.execute("SELECT COUNT(*), COALESCE(SUM(size),0) FROM backups").fetchone()
    files_disk = sum(p.stat().st_size for p in FILES_DIR.rglob("*") if p.is_file())
    return jsonify(
        users_total=users_total, users_active=users_active, users_with_state=states,
        attachments_count=atts[0], attachments_bytes=atts[1],
        backups_count=bks[0], backups_bytes=bks[1],
        files_on_disk_bytes=files_disk,
    )


# ---------- static (opt-in for single-container deploys) ----------

if STATIC_DIR:
    from flask import send_from_directory

    _static_root = Path(STATIC_DIR).resolve()

    @APP.route("/", defaults={"path": ""})
    @APP.route("/<path:path>")
    def serve_static(path):
        # /api/* must never be served from disk; if a real route exists Flask
        # will dispatch first, but unmatched /api/<...> falls through here and
        # we want a JSON 404, not an attempt to read api/<...> off disk.
        if path.startswith("api/") or path.startswith("api"):
            abort(404)
        if not path:
            path = "index.html"
        target = (_static_root / path).resolve()
        if _static_root not in target.parents and target != _static_root:
            abort(404)
        if not target.is_file():
            abort(404)
        return send_from_directory(_static_root, path)


init()

if __name__ == "__main__":
    APP.run(host="127.0.0.1", port=5050)
