"""Note<->project linking and the XLSX export endpoint."""
import io


def test_note_project_id_create_filter_patch(client):
    # Create a note linked to a board project
    r = client.post("/api/notes", json={"title": "site A note", "project_id": "p-fiber-1"})
    assert r.status_code == 201, r.data
    nid = r.json["id"]
    assert r.json["project_id"] == "p-fiber-1"

    # An unlinked note for contrast
    client.post("/api/notes", json={"title": "loose note"})

    # Filter by project returns only the linked one
    r = client.get("/api/notes?project_id=p-fiber-1")
    assert r.status_code == 200
    assert [n["id"] for n in r.json] == [nid]

    # Re-link to another project
    r = client.patch(f"/api/notes/{nid}", json={"project_id": "p-fiber-2"})
    assert r.status_code == 200
    assert r.json["project_id"] == "p-fiber-2"
    assert client.get("/api/notes?project_id=p-fiber-1").json == []

    # Unlink
    r = client.patch(f"/api/notes/{nid}", json={"project_id": ""})
    assert r.status_code == 200
    assert r.json["project_id"] is None


def test_export_xlsx_round_trip(client):
    payload = {
        "filename": "Fiber Project export",
        "sheets": [
            {"name": "Tasks", "headers": ["Title", "Done", "Priority"],
             "rows": [["Pull cable", "no", "high"], ["Terminate", "yes", "med"]]},
            {"name": "Notes", "headers": ["Title", "Date"], "rows": [["Site survey", "2026-06-11"]]},
        ],
    }
    r = client.post("/api/export/xlsx", json=payload)
    assert r.status_code == 200, r.data
    assert "spreadsheetml" in r.headers["Content-Type"]
    assert "Fiber_Project_export.xlsx" in r.headers.get("Content-Disposition", "")

    # It's a real workbook we can read back
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ["Tasks", "Notes"]
    ws = wb["Tasks"]
    assert [c.value for c in ws[1]] == ["Title", "Done", "Priority"]
    assert ws["A2"].value == "Pull cable"


def test_note_task_id_create_filter_patch(client):
    r = client.post("/api/notes", json={"title": "task note", "project_id": "p1", "task_id": "t-99"})
    assert r.status_code == 201, r.data
    nid = r.json["id"]
    assert r.json["task_id"] == "t-99"
    assert r.json["project_id"] == "p1"

    client.post("/api/notes", json={"title": "other"})
    r = client.get("/api/notes?task_id=t-99")
    assert [n["id"] for n in r.json] == [nid]

    # Unlink the task
    r = client.patch(f"/api/notes/{nid}", json={"task_id": ""})
    assert r.status_code == 200
    assert r.json["task_id"] is None
    assert client.get("/api/notes?task_id=t-99").json == []


def test_export_xlsx_neutralizes_formula_injection(client):
    payload = {"filename": "x", "sheets": [{"name": "S", "headers": ["A"],
               "rows": [["=1+2"], ["=HYPERLINK(\"http://evil\")"], ["+cmd"], ["@SUM(1)"], ["-1"], ["safe"]]}]}
    r = client.post("/api/export/xlsx", json=payload)
    assert r.status_code == 200
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.data))["S"]
    # No cell may be stored as a formula; values preserved verbatim as text.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            assert cell.data_type != "f", f"formula leaked: {cell.value!r}"
    assert ws["A2"].value == "=1+2"  # preserved, but as text


def test_export_xlsx_requires_auth(anon_client):
    assert anon_client.post("/api/export/xlsx", json={"sheets": []}).status_code == 401
